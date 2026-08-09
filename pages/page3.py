import streamlit as st
import pandas as pd
import numpy as np
from scipy.optimize import curve_fit
import matplotlib.pyplot as plt
import io
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# --- 页面配置 ---
st.set_page_config(layout="wide")
st.title("🧪 EC50 剂量-反应曲线分析工具")

# 创建两个标签页
tab1, tab2 = st.tabs(["📊 单组分析 (表格输入)", "📂 批量分析 (Excel导入)"])

# --- 核心算法：四参数逻辑斯蒂模型 (4PL) ---
def four_param_logistic(x, a, b, c, d):
    return d + (a - d) / (1 + 10**((c - x) * b))

# ============================
# TAB 1: 单组分析
# ============================
with tab1:
    # 初始化 Session State
    if 'single_data' not in st.session_state:
        st.session_state.single_data = pd.DataFrame({
            "Conc": [0.001, 0.01, 0.1, 1.0, 10.0, 100.0],
            "Rep1": [5.0, 12.0, 48.0, 82.0, 94.0, 95.0],
            "Rep2": [6.0, 10.0, 50.0, 80.0, 95.0, 96.0]
        })

    st.markdown("### 📝 数据输入与拟合结果")
    st.info("说明：左侧为原始数据输入，右侧为生成的曲线与参数。")

    # === 布局：强制左右分栏 (已调整为 1:1) ===
    col_left, col_right = st.columns([0.5, 0.5])

    with col_left:
        # 左侧：原始数据输入
        edited_df = st.data_editor(
            st.session_state.single_data,
            num_rows="dynamic",
            use_container_width=True,
            height=500,
            column_config={
                "Conc": st.column_config.NumberColumn("Conc (浓度)", format="%.3f"),
                "Rep1": st.column_config.NumberColumn("Rep1 (促进率%)", format="%.1f"),
                "Rep2": st.column_config.NumberColumn("Rep2 (促进率%)", format="%.1f"),
            }
        )

    with col_right:
        # 右侧：上方放图，下方放参数表
        if not edited_df.empty:
            try:
                # 1. 提取数据
                conc = edited_df["Conc"].values
                rep1 = edited_df["Rep1"].values
                rep2 = edited_df["Rep2"].values

                # 2. 计算平均值用于拟合
                mean_effect = (rep1 + rep2) / 2.0

                # 3. 过滤掉非数值或NaN
                valid_idx = ~np.isnan(conc) & ~np.isnan(mean_effect)
                x_data = conc[valid_idx]
                y_data = mean_effect[valid_idx]

                if len(x_data) < 3:
                    st.warning("数据点不足，请至少输入3行有效数据。")
                else:
                    # 4. 转换为 Log10
                    log_x = np.log10(x_data)

                    # 5. 拟合 4PL
                    try:
                        popt, pcov = curve_fit(four_param_logistic, log_x, y_data, p0=[100, 1, 0, 0], maxfev=10000)
                        a, b, c, d = popt
                        # Top, Hill, LogEC50, Bottom
                        ec50_val = 10**c
                        fit_success = True
                    except Exception:
                        fit_success = False

                    # 6. 右侧上部：绘图
                    st.markdown("#### 📈 剂量-反应曲线")
                    # 生成平滑曲线数据
                    x_plot = np.linspace(min(log_x)-1, max(log_x)+1, 200)
                    # 设置绘图风格
                    plt.style.use('default')
                    fig, ax = plt.subplots(figsize=(7, 5))

                    # 绘制散点 (蓝色)
                    ax.scatter(log_x, y_data, color='#1f77b4', edgecolors='white', s=60, label='Data Points', zorder=5)

                    if fit_success:
                        y_plot = four_param_logistic(x_plot, *popt)
                        # 绘制曲线 (黑色)
                        ax.plot(x_plot, y_plot, color='black', linewidth=2.5, label=f'4PL Fit')
                        ax.axvline(x=c, color='gray', linestyle='--', alpha=0.6)
                        ax.set_title(f'EC50 = {ec50_val:.4f}', fontsize=14, color='black', pad=20)
                    else:
                        st.warning("拟合失败，请检查数据是否呈现 S 型趋势。")

                    # 通用设置
                    ax.set_xlabel('Log10(Concentration)', fontsize=12)
                    ax.set_ylabel('Effect (%)', fontsize=12)
                    ax.legend(frameon=True, facecolor='white', framealpha=1)
                    ax.grid(True, linestyle='--', alpha=0.5, color='gray')

                    # 在 Streamlit 中显示
                    st.pyplot(fig)

                    # 7. 矢量图下载功能 (SVG格式)
                    if fit_success:
                        # 将图保存为 SVG (矢量图)
                        svg_io = io.StringIO()
                        fig.savefig(svg_io, format='svg', bbox_inches='tight')
                        svg_content = svg_io.getvalue()
                        svg_io.close()

                        # 创建下载按钮
                        b64 = base64.b64encode(svg_content.encode()).decode()
                        href = f'<a href="data:image/svg+xml;base64,{b64}" download="EC50_Curve.svg"><button style="margin-top:10px; padding:5px 10px; background-color:#007bff; color:white; border:none; border-radius:4px; cursor:pointer;">📥 下载矢量图 (SVG)</button></a>'
                        st.markdown(href, unsafe_allow_html=True)
                        st.caption("提示：SVG 格式可拖入 PPT 或 Excel，右键“取消组合”后可编辑颜色和线条。")

                    # 8. 右侧下部：参数表
                    if fit_success:
                        st.markdown("#### 📊 拟合参数 (ABCD)")
                        param_df = pd.DataFrame({
                            "Parameter": ["A (Top)", "B (Hill Slope)", "C (LogEC50)", "D (Bottom)", "EC50"],
                            "Value": [f"{a:.4f}", f"{b:.4f}", f"{c:.4f}", f"{d:.4f}", f"{ec50_val:.4f}"]
                        })
                        st.table(param_df)

            except Exception as e:
                st.error(f"计算出错: {str(e)}")
        else:
            st.info("请在左侧输入数据...")

# ============================
# TAB 2: 批量分析 (Excel导入 - 生成Excel报告版)
# ============================
with tab2:
    st.markdown("### 📂 上传 Excel 文件")
    st.info("格式要求：每三列一组（浓度, Rep1, Rep2）。支持 `Compound_A_Conc`, `Compound_A_eff`, `Unnamed` 格式。")
    uploaded_file = st.file_uploader("请选择 Excel 文件 (.xlsx)", type=["xlsx"])

    if uploaded_file:
        try:
            # 读取 Excel
            raw_df = pd.read_excel(uploaded_file)

            # === 核心修复：智能列名清洗 ===
            temp_cols = list(raw_df.columns)
            conc_indices = [i for i, col in enumerate(temp_cols) if 'Conc' in str(col)]
            final_col_map = {}

            for idx in conc_indices:
                final_col_map[idx] = temp_cols[idx]
                if idx + 1 < len(temp_cols):
                    prefix = str(final_col_map[idx]).split('_Conc')[0]
                    new_name_1 = f"{prefix}_eff_repeat1"
                    final_col_map[idx+1] = new_name_1
                if idx + 2 < len(temp_cols):
                    new_name_2 = f"{prefix}_eff_repeat2"
                    final_col_map[idx+2] = new_name_2

            renamed_cols = []
            for i, col in enumerate(raw_df.columns):
                if i in final_col_map:
                    renamed_cols.append(final_col_map[i])
                else:
                    renamed_cols.append(col)
            raw_df.columns = renamed_cols

            # 显示预览
            st.markdown("#### 数据预览 (已自动修正列名)")
            st.dataframe(raw_df.head(), use_container_width=True)

            # === 解析数据并分组 ===
            compounds_data = {}
            for col in raw_df.columns:
                if 'Conc' in str(col):
                    prefix = str(col).split('_Conc')[0]
                    conc_col = col
                    rep1_col = f"{prefix}_eff_repeat1"
                    rep2_col = f"{prefix}_eff_repeat2"
                    if rep1_col in raw_df.columns and rep2_col in raw_df.columns:
                        compounds_data[prefix] = {
                            'conc': raw_df[conc_col].values,
                            'rep1': raw_df[rep1_col].values,
                            'rep2': raw_df[rep2_col].values
                        }

            if not compounds_data:
                st.warning("未检测到符合格式的数据列（需包含 Conc, eff_repeat1, eff_repeat2）。")
            else:
                st.success(f"成功识别到 {len(compounds_data)} 个化合物数据。")
                results_summary = []
                all_charts_data = {} # 存储图表和数据的字典

                st.markdown("#### 📊 批量拟合结果")

                # === 布局：一行三个 ===
                compound_list = list(compounds_data.keys())

                # 循环处理每个化合物
                for i in range(0, len(compound_list), 3):
                    cols = st.columns(3)
                    for j in range(3):
                        index = i + j
                        if index < len(compound_list):
                            key = compound_list[index]
                            data = compounds_data[key]
                            x_raw = data['conc']
                            y1 = data['rep1']
                            y2 = data['rep2']
                            y_mean = (y1 + y2) / 2.0
                            valid = ~np.isnan(x_raw) & ~np.isnan(y_mean)
                            x = x_raw[valid]
                            y = y_mean[valid]

                            if len(x) < 3:
                                with cols[j]:
                                    st.warning(f"{key}: 数据点太少")
                                continue

                            log_x = np.log10(x)
                            try:
                                popt, _ = curve_fit(four_param_logistic, log_x, y, p0=[100, 1, 0, 0], maxfev=5000)
                                a, b, c, d = popt
                                ec50 = 10**c

                                # === 绘图 ===
                                plt.style.use('default')
                                fig, ax = plt.subplots(figsize=(6.5, 3.5)) # 调整了图表宽度
                                ax.scatter(log_x, y, color='#1f77b4', edgecolors='white', s=30, zorder=5)
                                x_plot = np.linspace(min(log_x)-0.5, max(log_x)+0.5, 100)
                                y_plot = four_param_logistic(x_plot, *popt)
                                ax.plot(x_plot, y_plot, color='black', linewidth=1.5)
                                ax.set_title(f"{key}\nEC50: {ec50:.3f}", fontsize=10, color='black', pad=10)
                                ax.set_xlabel("Log10(Conc)", fontsize=8)
                                ax.set_ylabel("Effect (%)", fontsize=8)
                                ax.tick_params(axis='both', which='major', labelsize=6)
                                ax.grid(True, linestyle='--', alpha=0.5, color='gray')

                                with cols[j]:
                                    st.markdown("<br>", unsafe_allow_html=True)
                                    st.pyplot(fig)

                                # 记录结果
                                results_summary.append({
                                    "Compound": key,
                                    "EC50": round(ec50, 4),
                                    "A (Top)": round(a, 2),
                                    "B (Hill)": round(b, 2),
                                    "C (LogEC50)": round(c, 2),
                                    "D (Bottom)": round(d, 2)
                                })

                                # 存储图表数据用于生成报告
                                img_io = io.BytesIO()
                                fig.savefig(img_io, format='png', bbox_inches='tight', dpi=300)
                                img_io.seek(0)
                                all_charts_data[key] = {
                                    'image_data': img_io.getvalue(),
                                    'params': results_summary[-1]
                                }
                                img_io.close()
                                plt.close(fig) # 关闭图表释放内存

                            except Exception as e:
                                with cols[j]:
                                    st.warning(f"{key}: 拟合失败")

                # ============================
                # === 生成并下载Excel报告 (已优化：3x2 矩阵布局 + 图下参数) ===
                # ============================
                if all_charts_data:
                    st.markdown("---")
                    st.markdown("### 🚀 导出Excel报告")

                    # 在内存中创建Excel文件
                    output = io.BytesIO()
                    wb = Workbook()
                    ws_summary = wb.active
                    ws_summary.title = "结果汇总"

                    # 写入汇总表头
                    summary_headers = ["化合物名称", "EC50", "A (Top)", "B (Hill)", "C (LogEC50)", "D (Bottom)"]
                    ws_summary.append(summary_headers)

                    # === 优化开始：创建一个专门的Sheet来放所有图表 (3x2 布局) ===
                    ws_charts = wb.create_sheet(title="批量分析结果")

                    # 设置初始行号
                    current_row = 1
                    items_in_current_sheet = 0

                    # 遍历所有化合物数据
                    for i, (compound_name, data) in enumerate(all_charts_data.items()):
                        # 检查是否需要换页 (每页6个图，即2行x3列)
                        if i > 0 and i % 6 == 0:
                            ws_charts = wb.create_sheet(title=f"批量分析结果_{(i//6)+1}")
                            current_row = 1
                            items_in_current_sheet = 0

                        params = data['params']

                        # 计算当前图应该在哪一列 (0, 1, 2 对应 A, D, G 列)
                        # 我们让每列之间空一列，所以列号是 1, 4, 7
                        col_offset = items_in_current_sheet % 3
                        excel_col = 1 + (col_offset * 3) # A列(1), D列(4), G列(7)

                        # 计算当前图应该在哪一行 (0-2 在第一行, 3-5 在第二行)
                        # 每个图占大约 10 行高度 (2行参数 + 1行空 + 6行图 + 1行空)
                        row_offset = (items_in_current_sheet // 3) * 12

                        # 1. 写入参数 (在图片上方)
                        # 标题
                        ws_charts.cell(row=current_row + row_offset, column=excel_col, value=f"--- {compound_name} ---")
                        # 表头
                        ws_charts.cell(row=current_row + 1 + row_offset, column=excel_col, value="参数")
                        ws_charts.cell(row=current_row + 1 + row_offset, column=excel_col+1, value="数值")
                        # 参数值
                        param_keys = ["EC50", "A (Top)", "B (Hill)", "C (LogEC50)", "D (Bottom)"]
                        for idx, k in enumerate(param_keys):
                            ws_charts.cell(row=current_row + 2 + idx + row_offset, column=excel_col, value=k)
                            ws_charts.cell(row=current_row + 2 + idx + row_offset, column=excel_col+1, value=params[k])

                        # 2. 插入图片 (在参数下方)
                        # 图片起始行 = 当前块起始行 + 参数行数(2+5=7) + 1行间距
                        img_start_row = current_row + 8 + row_offset
                        img_data = io.BytesIO(data['image_data'])
                        img = OpenpyxlImage(img_data)
                        # 调整图片大小以适应单元格 (宽约3列，高约6行)
                        img.width = 380
                        img.height = 260
                        # 图片插入位置
                        cell_coordinate = f'{ws_charts.cell(row=img_start_row, column=excel_col).column_letter}{img_start_row}'
                        ws_charts.add_image(img, cell_coordinate)

                        # 3. 在汇总表中添加一行
                        ws_summary.append([params["Compound"], params["EC50"], params["A (Top)"], params["B (Hill)"], params["C (LogEC50)"], params["D (Bottom)"]])

                        items_in_current_sheet += 1

                    # 调整汇总表列宽
                    for col in ws_summary.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws_summary.column_dimensions[column].width = adjusted_width

                    # 调整图表Sheet的列宽，让图片显示更完整
                    # 设置 A, D, G 列宽
                    for col_letter in ['A', 'D', 'G']:
                        ws_charts.column_dimensions[col_letter].width = 55

                    wb.save(output)
                    output.seek(0)

                    st.download_button(
                        label="📥 下载完整Excel报告 (3x2矩阵版)",
                        data=output,
                        file_name="EC50_批量分析报告_3x2矩阵.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.caption("点击下载，你将获得一个包含汇总表和所有图表的Excel文件（每页6图，3x2排列，参数在图上方）。")

                # === 底部汇总表 ===
                if results_summary:
                    st.markdown("---")
                    st.markdown("### 📑 批量计算结果汇总")
                    summary_df = pd.DataFrame(results_summary)
                    csv = summary_df.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        label="📥 下载结果为 CSV",
                        data=csv,
                        file_name='EC50_Batch_Results.csv',
                        mime='text/csv',
                    )
                    st.dataframe(summary_df, use_container_width=True)

        except Exception as e:
            st.error(f"处理文件出错: {str(e)}")
            st.exception(e)
